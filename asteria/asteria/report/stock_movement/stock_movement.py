# Copyright (c) 2026, Viral and contributors
# For license information, please see license.txt

import json
from io import BytesIO

import frappe
from frappe import _
from frappe.utils import flt, getdate, today, date_diff


# Voucher types / Stock Entry purposes that represent a genuine inward procurement
INWARD_VOUCHER_TYPES = {"Purchase Receipt", "Purchase Invoice"}
INWARD_SE_PURPOSES = {"Material Receipt", "Opening Stock"}


def execute(filters=None):
	if not filters:
		filters = {}
	columns = get_columns()
	data = get_data(filters)
	return columns, data


def get_columns():
	return [
		{"label": _("Item Code"), "fieldname": "item_code", "fieldtype": "Link", "options": "Item", "width": 120},
		{"label": _("Item Name"), "fieldname": "item_name", "fieldtype": "Data", "width": 150},
		{"label": _("Item Group"), "fieldname": "item_group", "fieldtype": "Data", "width": 110},
		{"label": _("Batch/SRL"), "fieldname": "batch_serial_type", "fieldtype": "Data", "width": 80},
		{"label": _("Serial/Batch No"), "fieldname": "serial_batch_no", "fieldtype": "Data", "width": 130},
		{"label": _("Txn Type"), "fieldname": "txn_type", "fieldtype": "Data", "width": 110},
		# IN columns
		{"label": _("IN: From Warehouse"), "fieldname": "in_from_warehouse", "fieldtype": "Link", "options": "Warehouse", "width": 150},
		{"label": _("IN: To Warehouse"), "fieldname": "in_to_warehouse", "fieldtype": "Link", "options": "Warehouse", "width": 150},
		{"label": _("IN: Voucher Type"), "fieldname": "in_voucher_type", "fieldtype": "Data", "width": 140},
		{"label": _("IN: Voucher No"), "fieldname": "in_voucher_no", "fieldtype": "Dynamic Link", "options": "in_voucher_type", "width": 150},
		{"label": _("IN: Date"), "fieldname": "in_date", "fieldtype": "Date", "width": 100},
		{"label": _("IN: Qty"), "fieldname": "in_qty", "fieldtype": "Float", "precision": 2, "width": 80},
		{"label": _("UoM"), "fieldname": "uom", "fieldtype": "Data", "width": 60},
		{"label": _("IN: Rate"), "fieldname": "in_rate", "fieldtype": "Currency", "precision": 2, "width": 100},
		{"label": _("IN: Amount"), "fieldname": "in_amount", "fieldtype": "Currency", "precision": 2, "width": 110},
		# OUT columns (all movements after inward)
		{"label": _("OUT: Date"), "fieldname": "out_date", "fieldtype": "Date", "width": 100},
		{"label": _("OUT: From Warehouse"), "fieldname": "out_from_warehouse", "fieldtype": "Link", "options": "Warehouse", "width": 150},
		{"label": _("OUT: To Warehouse"), "fieldname": "out_to_warehouse", "fieldtype": "Link", "options": "Warehouse", "width": 150},
		{"label": _("OUT: Voucher Type"), "fieldname": "out_voucher_type", "fieldtype": "Data", "width": 140},
		{"label": _("OUT: Voucher No"), "fieldname": "out_voucher_no", "fieldtype": "Dynamic Link", "options": "out_voucher_type", "width": 150},
		{"label": _("OUT: Consumption Date"), "fieldname": "out_consumption_date", "fieldtype": "Date", "width": 120},
		{"label": _("OUT: Qty"), "fieldname": "out_qty", "fieldtype": "Float", "precision": 2, "width": 80},
		{"label": _("OUT: UoM"), "fieldname": "out_uom", "fieldtype": "Data", "width": 60},
		{"label": _("OUT: Rate"), "fieldname": "out_rate", "fieldtype": "Currency", "precision": 2, "width": 100},
		{"label": _("OUT: Amount"), "fieldname": "out_amount", "fieldtype": "Currency", "precision": 2, "width": 110},
		# Summary columns
		{"label": _("Inward Qty"), "fieldname": "inward_qty", "fieldtype": "Float", "precision": 2, "width": 100},
		{"label": _("Inward Value"), "fieldname": "inward_value", "fieldtype": "Currency", "precision": 2, "width": 110},
		{"label": _("Outward Qty"), "fieldname": "outward_qty", "fieldtype": "Float", "precision": 2, "width": 100},
		{"label": _("Outward Value"), "fieldname": "outward_value", "fieldtype": "Currency", "precision": 2, "width": 110},
		{"label": _("Closing Qty"), "fieldname": "closing_qty", "fieldtype": "Float", "precision": 2, "width": 100},
		{"label": _("Closing Value"), "fieldname": "closing_value", "fieldtype": "Currency", "precision": 2, "width": 110},
		{"label": _("Warehouse"), "fieldname": "warehouse", "fieldtype": "Data", "width": 180},
		{"label": _("Ageing (Days)"), "fieldname": "ageing_days", "fieldtype": "Int", "width": 110},
	]


def get_qualifying_keys(filters):
	"""
	Step 1 — fast, narrow query.
	Find batch/serial nos (or item+warehouse for plain items) that had a genuine
	inward entry (Purchase Receipt / Material Receipt) within the date range.
	Date filter is applied HERE so the SQL is small and uses indexed columns.
	"""
	conditions = [
		"sle.docstatus = 1",
		"sle.is_cancelled = 0",
		"sle.actual_qty > 0",
		"(sle.voucher_type IN ('Purchase Receipt', 'Purchase Invoice') OR "
		"(sle.voucher_type = 'Stock Entry' AND se.purpose IN ('Material Receipt', 'Opening Stock')))",
	]
	if filters.get("from_date"):
		conditions.append("sle.posting_date >= %(from_date)s")
	if filters.get("to_date"):
		conditions.append("sle.posting_date <= %(to_date)s")
	if filters.get("company"):
		conditions.append("sle.company = %(company)s")
	if filters.get("item_code"):
		conditions.append("sle.item_code = %(item_code)s")
	if filters.get("item_group"):
		conditions.append("item.item_group = %(item_group)s")
	if filters.get("warehouse"):
		conditions.append("sle.warehouse = %(warehouse)s")
	if filters.get("batch_no"):
		conditions.append(
			"(sle.batch_no = %(batch_no)s OR EXISTS ("
			"SELECT 1 FROM `tabSerial and Batch Entry` sbe "
			"WHERE sbe.parent = sle.serial_and_batch_bundle AND sbe.batch_no = %(batch_no)s))"
		)
	if filters.get("serial_no"):
		conditions.append(
			"(sle.serial_no = %(serial_no)s OR EXISTS ("
			"SELECT 1 FROM `tabSerial and Batch Entry` sbe "
			"WHERE sbe.parent = sle.serial_and_batch_bundle AND sbe.serial_no = %(serial_no)s))"
		)

	where = " AND ".join(conditions)
	rows = frappe.db.sql(
		f"""
		SELECT DISTINCT
			sle.item_code,
			item.has_batch_no,
			item.has_serial_no,
			sle.batch_no,
			sle.serial_no,
			sle.serial_and_batch_bundle,
			sle.warehouse
		FROM `tabStock Ledger Entry` sle
		LEFT JOIN `tabItem` item ON item.name = sle.item_code
		LEFT JOIN `tabStock Entry` se
			ON se.name = sle.voucher_no AND sle.voucher_type = 'Stock Entry'
		WHERE {where}
		""",
		filters,
		as_dict=True,
	)

	batch_nos = set()
	serial_nos = set()
	plain_keys = set()  # (item_code, warehouse) for non-batch items

	# Collect all bundle names in one shot to avoid N+1 queries
	bundle_names = [r.serial_and_batch_bundle for r in rows if r.serial_and_batch_bundle]
	if bundle_names:
		escaped = ", ".join(frappe.db.escape(b) for b in bundle_names)
		bundle_entries = frappe.db.sql(
			f"SELECT batch_no, serial_no FROM `tabSerial and Batch Entry` WHERE parent IN ({escaped})",
			as_dict=True,
		)
		for e in bundle_entries:
			if e.batch_no:
				batch_nos.add(e.batch_no)
			if e.serial_no:
				serial_nos.add(e.serial_no)

	for r in rows:
		if r.has_batch_no or r.has_serial_no:
			if not r.serial_and_batch_bundle:
				if r.batch_no:
					batch_nos.add(r.batch_no)
				if r.serial_no:
					serial_nos.update(s.strip() for s in r.serial_no.split("\n") if s.strip())
		else:
			plain_keys.add((r.item_code, r.warehouse))

	return {"batch_nos": batch_nos, "serial_nos": serial_nos, "plain_keys": plain_keys}


def fetch_sle_for_keys(filters, qualifying):
	"""
	Step 2 — fetch ALL movements (no date limit) for the qualifying batches/items.
	The IN-list is built from the small set returned by get_qualifying_keys().
	"""
	batch_nos = qualifying["batch_nos"]
	serial_nos = qualifying["serial_nos"]
	plain_keys = qualifying["plain_keys"]

	if not batch_nos and not serial_nos and not plain_keys:
		return []

	base_conditions = ["sle.docstatus = 1", "sle.is_cancelled = 0"]
	if filters.get("company"):
		base_conditions.append("sle.company = %(company)s")

	key_parts = []
	if batch_nos:
		esc = ", ".join(frappe.db.escape(b) for b in batch_nos)
		key_parts.append(
			f"sle.batch_no IN ({esc})"
		)
		key_parts.append(
			f"EXISTS (SELECT 1 FROM `tabSerial and Batch Entry` sbe "
			f"WHERE sbe.parent = sle.serial_and_batch_bundle AND sbe.batch_no IN ({esc}))"
		)
	if serial_nos:
		esc = ", ".join(frappe.db.escape(s) for s in serial_nos)
		key_parts.append(f"sle.serial_no IN ({esc})")
		key_parts.append(
			f"EXISTS (SELECT 1 FROM `tabSerial and Batch Entry` sbe "
			f"WHERE sbe.parent = sle.serial_and_batch_bundle AND sbe.serial_no IN ({esc}))"
		)
	if plain_keys:
		pairs = " OR ".join(
			f"(sle.item_code = {frappe.db.escape(ic)} AND sle.warehouse = {frappe.db.escape(wh)})"
			for ic, wh in plain_keys
		)
		key_parts.append(f"({pairs})")

	where = " AND ".join(base_conditions) + f" AND ({' OR '.join(key_parts)})"

	return frappe.db.sql(
		f"""
		SELECT
			sle.name,
			sle.item_code,
			item.item_name,
			item.item_group,
			item.has_serial_no,
			item.has_batch_no,
			item.stock_uom,
			sle.warehouse,
			sle.actual_qty,
			sle.valuation_rate,
			sle.stock_value_difference,
			sle.voucher_type,
			sle.voucher_no,
			sle.posting_date,
			sle.batch_no,
			sle.serial_no,
			sle.serial_and_batch_bundle,
			se.purpose AS stock_entry_purpose
		FROM `tabStock Ledger Entry` sle
		LEFT JOIN `tabItem` item ON item.name = sle.item_code
		LEFT JOIN `tabStock Entry` se
			ON se.name = sle.voucher_no AND sle.voucher_type = 'Stock Entry'
		WHERE {where}
		ORDER BY sle.item_code, sle.batch_no, sle.serial_no, sle.posting_date, sle.posting_time
		""",
		filters,
		as_dict=True,
	)


def is_inward_txn(sle):
	"""
	Return True only for genuine procurement / receipt entries.
	Everything else (transfers, manufacture, delivery) is treated as a movement (OUT).
	"""
	if sle.voucher_type in INWARD_VOUCHER_TYPES:
		return True
	if sle.voucher_type == "Stock Entry" and (sle.get("stock_entry_purpose") or "") in INWARD_SE_PURPOSES:
		return True
	return False


def get_movement_label(sle):
	"""Human-readable movement type for OUT rows."""
	vt = sle.voucher_type
	purpose = sle.get("stock_entry_purpose") or ""
	if vt == "Stock Entry":
		label_map = {
			"Material Transfer": "Transfer",
			"Material Transfer for Manufacture": "Transfer",
			"Manufacture": "Manufacture",
			"Material Issue": "Issue",
			"Material Receipt": "Receipt",
			"Send to Subcontractor": "Subcontract",
		}
		return label_map.get(purpose, purpose or "Stock Entry")
	label_map = {
		"Delivery Note": "Delivery",
		"Sales Invoice": "Delivery",
		"Purchase Return": "Return",
		"Stock Reconciliation": "Reconciliation",
	}
	return label_map.get(vt, vt)


def get_serial_batch_entries(bundle_name):
	"""Return list of (serial_no, batch_no, qty) from a Serial and Batch Bundle."""
	return frappe.db.sql(
		"""
		SELECT serial_no, batch_no, qty
		FROM `tabSerial and Batch Entry`
		WHERE parent = %s
		ORDER BY idx
		""",
		bundle_name,
		as_dict=True,
	)


def make_base_row(group):
	# Initialize ALL column fieldnames to None so Frappe's export never hits a KeyError
	row = {c["fieldname"]: None for c in get_columns()}
	row.update({
		"item_code": group["item_code"],
		"item_name": group["item_name"],
		"item_group": group["item_group"],
		"batch_serial_type": group["batch_serial_type"],
		"serial_batch_no": group["serial_batch_no"],
		"uom": group["uom"],
	})
	return row


def init_group(sle, batch_serial_type, serial_batch_no):
	return {
		"item_code": sle.item_code,
		"item_name": sle.item_name,
		"item_group": sle.item_group,
		"batch_serial_type": batch_serial_type,
		"serial_batch_no": serial_batch_no,
		"uom": sle.stock_uom,
		"transactions": [],
		"inward_qty": 0.0,
		"inward_value": 0.0,
		"outward_qty": 0.0,
		"outward_value": 0.0,
		"warehouses": set(),
		"warehouse_data": {},   # per-warehouse breakdown: {wh: {inward_qty, inward_value, outward_qty, outward_value}}
		"first_inward_date": None,
	}


def add_txn_to_group(group, sle, qty, value, rate):
	inward = is_inward_txn(sle)
	group["transactions"].append({
		"sle": sle,
		"qty": qty,
		"value": value,
		"rate": rate,
		"is_inward": inward,
	})
	wh = sle.warehouse
	group["warehouses"].add(wh)

	# Per-warehouse breakdown
	if wh not in group["warehouse_data"]:
		group["warehouse_data"][wh] = {"inward_qty": 0.0, "inward_value": 0.0, "outward_qty": 0.0, "outward_value": 0.0}
	wh_data = group["warehouse_data"][wh]

	if qty > 0:
		# Positive qty = stock increased at this warehouse (purchase or transfer-in)
		group["inward_qty"] += qty
		group["inward_value"] += abs(value)
		wh_data["inward_qty"] += qty
		wh_data["inward_value"] += abs(value)
		# Ageing: track first genuine procurement date only
		if inward and not group["first_inward_date"]:
			group["first_inward_date"] = sle.posting_date
	else:
		# Negative qty = stock decreased (delivery, transfer-out, consumption)
		group["outward_qty"] += abs(qty)
		group["outward_value"] += abs(value)
		wh_data["outward_qty"] += abs(qty)
		wh_data["outward_value"] += abs(value)


def get_data(filters):
	# Step 1: narrow query — find which batches/serials had an inward in the date range
	qualifying = get_qualifying_keys(filters)

	# Step 2: fetch ALL movements for only those batches (no date restriction)
	sle_list = fetch_sle_for_keys(filters, qualifying)
	if not sle_list:
		return []

	# groups keyed by (item_code, serial_batch_no) — for plain items keyed by (item_code, warehouse)
	groups = {}

	for sle in sle_list:
		has_serial = sle.has_serial_no
		has_batch = sle.has_batch_no

		if has_serial or has_batch:
			batch_serial_type = "Serial" if has_serial else "Batch"

			if sle.serial_and_batch_bundle:
				# Newer ERPNext: unpack the bundle entries
				bundle_entries = get_serial_batch_entries(sle.serial_and_batch_bundle)
				if bundle_entries:
					total_bundle_qty = sum(flt(b.qty) for b in bundle_entries) or 1
					for entry in bundle_entries:
						sb_no = entry.serial_no or entry.batch_no or ""
						key = (sle.item_code, sb_no)
						if key not in groups:
							groups[key] = init_group(sle, batch_serial_type, sb_no)
						ratio = flt(entry.qty) / total_bundle_qty
						qty = flt(sle.actual_qty) * ratio
						value = flt(sle.stock_value_difference) * ratio
						add_txn_to_group(groups[key], sle, qty, value, flt(sle.valuation_rate))
					continue

			# Older ERPNext: batch_no / serial_no fields directly on SLE
			if has_serial and sle.serial_no:
				serial_numbers = [s.strip() for s in sle.serial_no.split("\n") if s.strip()]
			elif has_batch and sle.batch_no:
				serial_numbers = [sle.batch_no]
			else:
				serial_numbers = [""]

			n = len(serial_numbers) or 1
			for sb_no in serial_numbers:
				key = (sle.item_code, sb_no)
				if key not in groups:
					groups[key] = init_group(sle, batch_serial_type, sb_no)
				qty = flt(sle.actual_qty) / n
				value = flt(sle.stock_value_difference) / n
				add_txn_to_group(groups[key], sle, qty, value, flt(sle.valuation_rate))

		else:
			# Plain item — group by item + warehouse
			key = (sle.item_code, sle.warehouse)
			if key not in groups:
				groups[key] = init_group(sle, "", "")
				groups[key]["warehouses"].add(sle.warehouse)
			qty = flt(sle.actual_qty)
			value = flt(sle.stock_value_difference)
			add_txn_to_group(groups[key], sle, qty, value, flt(sle.valuation_rate))

	# Build report rows
	data = []
	today_date = getdate(today())

	for group in groups.values():
		inward_qty = group["inward_qty"]
		inward_value = group["inward_value"]
		outward_qty = group["outward_qty"]
		outward_value = group["outward_value"]
		closing_qty = inward_qty - outward_qty
		closing_value = inward_value - outward_value

		ageing_days = 0
		if group["first_inward_date"]:
			ageing_days = date_diff(today_date, group["first_inward_date"])

		for txn in group["transactions"]:
			sle = txn["sle"]
			qty = txn["qty"]
			value = txn["value"]
			rate = txn["rate"]
			inward = txn["is_inward"]

			row = make_base_row(group)

			if inward:
				# Purchase Receipt / Material Receipt → IN section
				row.update({
					"txn_type": "Inward",
					"in_to_warehouse": sle.warehouse,
					"in_voucher_type": sle.voucher_type,
					"in_voucher_no": sle.voucher_no,
					"in_date": sle.posting_date,
					"in_qty": abs(qty),
					"in_rate": rate,
					"in_amount": abs(value),
				})
			else:
				# All movements after inward (transfers, manufacture, delivery) → OUT section
				# For transfers: qty < 0 means stock left this warehouse (from),
				#                qty > 0 means stock arrived at this warehouse (to).
				out_from = sle.warehouse if qty < 0 else None
				out_to = sle.warehouse if qty > 0 else None
				row.update({
					"txn_type": get_movement_label(sle),
					"out_date": sle.posting_date,
					"out_from_warehouse": out_from,
					"out_to_warehouse": out_to,
					"out_voucher_type": sle.voucher_type,
					"out_voucher_no": sle.voucher_no,
					"out_consumption_date": sle.posting_date,
					"out_qty": abs(qty),
					"out_uom": group["uom"],
					"out_rate": rate,
					"out_amount": abs(value),
				})

			data.append(row)

		# One summary row per warehouse for this item/batch group
		for wh in sorted(group["warehouse_data"].keys()):
			wh_data = group["warehouse_data"][wh]
			wh_inward_qty   = wh_data["inward_qty"]
			wh_inward_value = wh_data["inward_value"]
			wh_outward_qty   = wh_data["outward_qty"]
			wh_outward_value = wh_data["outward_value"]
			wh_closing_qty   = wh_inward_qty - wh_outward_qty
			wh_closing_value = wh_inward_value - wh_outward_value

			summary = make_base_row(group)
			summary.update({
				"txn_type": "",
				"inward_qty": wh_inward_qty,
				"inward_value": wh_inward_value,
				"outward_qty": wh_outward_qty,
				"outward_value": wh_outward_value,
				"closing_qty": wh_closing_qty,
				"closing_value": wh_closing_value,
				"warehouse": wh,
				"ageing_days": ageing_days,
			})
			data.append(summary)

	return data


def _build_workbook(filters):
	"""Build and return an openpyxl Workbook for the Stock Movement report."""
	import openpyxl
	from openpyxl.styles import Alignment, Font, PatternFill
	from openpyxl.utils import get_column_letter

	COLORS = {
		"Inward": "D4EDDA",   # green
		"":       "E8F0FE",   # blue-grey  (summary row)
	}
	MOVEMENT_COLOR = "FFF3CD"   # amber — Transfer, Delivery, Manufacture …
	HEADER_BG      = "2F5496"
	HEADER_FG      = "FFFFFF"

	columns, data = execute(filters)
	fieldnames = [c["fieldname"] for c in columns]

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Stock Movement"
	ws.freeze_panes = "A2"

	header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
	header_font = Font(name="Calibri", bold=True, color=HEADER_FG)
	ws.append([c["label"] for c in columns])
	for cell in ws[1]:
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal="center", wrap_text=True)

	for i, col in enumerate(columns, 1):
		ws.column_dimensions[get_column_letter(i)].width = max(col.get("width", 100) / 7, 10)

	for row_data in data:
		txn  = row_data.get("txn_type", "")
		bg   = COLORS.get(txn, MOVEMENT_COLOR)
		fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
		font = Font(name="Calibri", bold=(txn == ""))

		ws.append([row_data.get(fn) for fn in fieldnames])
		for cell in ws[ws.max_row]:
			cell.fill = fill
			cell.font = font

	return wb


@frappe.whitelist()
def download_excel_with_colors(filters):
	if isinstance(filters, str):
		filters = json.loads(filters)

	wb = _build_workbook(filters)
	xlsx_file = BytesIO()
	wb.save(xlsx_file)

	frappe.response["filename"]    = "Stock_Movement.xlsx"
	frappe.response["filecontent"] = xlsx_file.getvalue()
	frappe.response["type"]        = "binary"


@frappe.whitelist()
def enqueue_excel_download(filters):
	"""Enqueue Excel generation as a background job and notify user when ready."""
	if isinstance(filters, str):
		filters = json.loads(filters)

	frappe.enqueue(
		"asteria.asteria.report.stock_movement.stock_movement.generate_and_save_excel",
		queue="long",
		timeout=1800,
		filters=filters,
		user=frappe.session.user,
	)


def generate_and_save_excel(filters, user):
	"""Background job: build Excel, save as a private File, then notify the user."""
	wb = _build_workbook(filters)

	xlsx_file = BytesIO()
	wb.save(xlsx_file)

	file_doc = frappe.get_doc({
		"doctype": "File",
		"file_name": "Stock_Movement.xlsx",
		"content": xlsx_file.getvalue(),
		"is_private": 1,
		"folder": "Home/Attachments",
	})
	file_doc.insert(ignore_permissions=True)
	frappe.db.commit()

	frappe.publish_realtime(
		"eval_js",
		(
			"frappe.msgprint("
			"'<a href=\"" + file_doc.file_url + "\" target=\"_blank\" download>"
			"Click here to download Stock Movement Excel</a>',"
			"'Excel Ready'"
			");"
		),
		user=user,
	)
